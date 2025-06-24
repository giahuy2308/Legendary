from ..models.base_models import Exam, MediaResource
from ..models.kd_models import KDQuestion
from ..models.vcnv_models import VCNVQuestion
from ..models.tt_models import TTQuestion
from ..models.vd_models import VDQuestion
from ..models.chp_models import CHPQuestion

from django.core.files.base import ContentFile
from django.conf import settings

import openpyxl 
import os

def create_media(question,media_file_name, round):
    media_path = os.path.join(DIR, round ,media_file_name)
    
    with open(media_path, "rb") as mf:
        media = MediaResource.objects.create(
            question=question,
            media_type=question.question_data_type,
        )
        media.resource.save(media_file_name, ContentFile(mf.read()))
        media.save()

def get_question(QuestionModel,exam,ws, step, max_row,target_player_no, **kwargs):
    for row in range(1 + step, max_row + step + 1):
        question_text = ws.cell(row=row, column=2).value
        answer_text = ws.cell(row=row, column=3).value
        image = ws.cell(row=row, column=4).value
        audio = ws.cell(row=row, column=5).value
        video = ws.cell(row=row, column=6).value

        question = QuestionModel.objects.create(
            question_no=row-step,
            exam=exam,
            question_text=question_text,
            answer_text=answer_text,
            question_data_type="Text",
            question_value=10,
            target_player_no=target_player_no,
        )

        if ws.title in ("khởi động","về đích"): 
            if ws.title == "về đích":
                value = ws.cell(row=row, column=1).value
                question.question_value= value

            if image:
                question.question_data_type = "Image"
                create_media(question,image,ws.title)
            elif audio:
                question.question_data_type = "Audio"
                create_media(question,audio,ws.title)
            elif video:
                question.question_data_type = "Video"
                create_media(question,video,ws.title)

        elif ws.title == "vượt chướng ngại vật":
            if image :
                if answer_text:
                    question.question_data_type = "Audio"
                else:
                    question.question_data_type = "Image"
                create_media(question,image,ws.title)

        elif ws.title == "tăng tốc":
            if audio:
                question.question_data_type = "Image"
                create_media(question,image, ws.title)
                create_media(question,audio, ws.title)
            else:
                question.question_data_type = "Video"
                create_media(question,image, ws.title)
        
        for update_field, update_value in kwargs.items():
            setattr(question, update_field, update_value)

        question.save()

def import_kd(kd_ws, exam):
    # Phần thi cá nhân
    for player_no in range(1,5):
        step = (player_no-1) * 6 + 4
        get_question(KDQuestion, exam, kd_ws, step, 6, player_no)
        
    # Phần thi đối kháng
    step = 30
    get_question(KDQuestion, exam, kd_ws, step, 12, 0)

def import_vcnv(vcnv_ws,exam):
    step = 2 
    get_question(VCNVQuestion, exam, vcnv_ws, step, 1, 0, question_type="CNV", picture_opened=True, question_shown=False, question_no=0)
    step = 4
    get_question(VCNVQuestion, exam, vcnv_ws, step, 4, 0, question_type="HN", picture_opened=False, question_shown=False)
    step = 8
    get_question(VCNVQuestion, exam, vcnv_ws, step, 1, 0, question_type="TT", picture_opened=False, question_shown=False, question_no=5)

def import_tt(tt_ws, exam):
    step = 3
    get_question(TTQuestion, exam, tt_ws, step, 4, 0, question_value=40)

def import_vd(vd_ws, exam):
    for player_no in range(1,5):
        step = 4 + (player_no-1) * 8
        get_question(VDQuestion, exam, vd_ws, step, 6, player_no)

def import_chp(chp_ws,exam):
    step = 3 
    get_question(CHPQuestion, exam, chp_ws, step, 3, 0)


def import_exam(exam_dir):
    # Get exam's absolute dir
    global DIR 
    DIR = os.path.join(settings.MEDIA_ROOT, "Exam", exam_dir)

    # Get excel file path
    for file in os.listdir(DIR):
        if file.endswith(".xlsx"):
            exam_file = file
    file_path = os.path.join(DIR, exam_file)

    # WorkBook
    wb = openpyxl.load_workbook(file_path) 

    # Worksheet
    kd_ws = wb["khởi động"]
    vcnv_ws = wb["vượt chướng ngại vật"]
    tt_ws = wb["tăng tốc"]
    vd_ws = wb["về đích"]
    chp_ws = wb["câu hỏi phụ"]

    # Create Exam
    exam = Exam.objects.create(title=exam_dir)

    # Import data each round
    import_kd(kd_ws,exam)
    import_vcnv(vcnv_ws, exam)
    import_tt(tt_ws, exam)
    import_vd(vd_ws, exam)
    import_chp(chp_ws, exam)
    
    exam.save()